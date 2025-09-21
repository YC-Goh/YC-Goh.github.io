
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from typing import Union, Tuple
from PIL import Image, ImageDraw, ImageFont
from typing import Union, Tuple

LATEST_GAME_VERSION = "1.3.4"
LATEST_FILTER_VERSION = "0"
RULE_OPTIONS = ["Rule Description", "Emphasise", "Recolour", "Sound", "Beam Colour", "Acolyte", "Primalist", "Mage", "Rogue", "Sentinel"]
RARITY_ORDER = ["NORMAL", "MAGIC", "RARE", "EXALTED", "UNIQUE", "LEGENDARY", "SET"]
SLOT_ORDER = [
    "One-Handed Axe", "One-Handed Mace", "One-Handed Sword", "Sceptre", "Dagger", "Wand", 
    "Two-Handed Axe", "Two-Handed Mace", "Two-Handed Sword", "Two-Handed Spear", "Two-Handed Staff", "Bow", 
    "Off-Hand Catalyst", "Shield", "Quiver", 
    "Helmet", "Body Armor", "Boots", "Gloves", "Belt", 
    "Ring", "Amulet", "Relic", 
    "Small Idol", "Minor Idol", "Humble Idol", "Stout Idol", 
    "Huge Idol", "Grand Idol", "Large Idol", "Ornate Idol", 
    "Adorned Idol", 
]
SLOT_ORDER_WITH_PRIMORDIAL = [
    slot for baseslot in SLOT_ORDER for slot in ([baseslot, f"Primordial {baseslot}",] if not baseslot.endswith(("Idol", )) else [baseslot,])
]
META_SLOT_SELECTOR = {
    "All One-Handed Weapons": [
        "One-Handed Axe", "One-Handed Mace", "Sceptre", "One-Handed Sword", "Wand", "Dagger", 
    ], 
    "All Two-Handed Weapons": [
        "Two-Handed Axe", "Two-Handed Mace", "Two-Handed Sword", "Two-Handed Spear", "Two-Handed Staff", "Bow", 
    ], 
    "All Weapons": [
        "One-Handed Axe", "One-Handed Mace", "Sceptre", "One-Handed Sword", "Wand", "Dagger", 
        "Two-Handed Axe", "Two-Handed Mace", "Two-Handed Sword", "Two-Handed Spear", "Two-Handed Staff", "Bow", 
    ], 
    "All Off-Hand": [
        "Off-Hand Catalyst", "Shield", "Quiver", 
    ], 
    "All Armour": [
        "Helmet", "Body Armor", "Belt", "Boots", "Gloves", 
    ], 
    "All Jewellery": [
        "Amulet", "Ring", "Relic", 
    ], 
    "All Equipment": [
        "Off-Hand Catalyst", "Shield", "Quiver", 
        "Helmet", "Body Armor", "Belt", "Boots", "Gloves", 
        "Amulet", "Ring", "Relic", 
    ], 
    "All Generic Idols": [
        "Small Idol", "Minor Idol", "Humble Idol", "Stout Idol", 
    ], 
    "All Class-Specific Idols": [
        "Grand Idol", "Huge Idol", "Ornate Idol", "Large Idol", "Adorned Idol", 
    ], 
    "All Idols": [
        "Small Idol", "Minor Idol", "Humble Idol", "Stout Idol", 
        "Grand Idol", "Huge Idol", "Ornate Idol", "Large Idol", "Adorned Idol", 
    ], 
}
SLOT_TO_TYPE_MAP = {
    "One-Handed Axe": ["ONE_HANDED_AXE", ], 
    "One-Handed Mace": ["ONE_HANDED_MACES", ], 
    "Sceptre": ["ONE_HANDED_SCEPTRE", ], 
    "One-Handed Sword": ["ONE_HANDED_SWORD", ], 
    "Wand": ["WAND", ], 
    "Dagger": ["ONE_HANDED_DAGGER", ], 
    "All One-Handed Weapons": [
        "ONE_HANDED_AXE", "ONE_HANDED_MACES", "ONE_HANDED_SCEPTRE", "ONE_HANDED_SWORD", "WAND", "ONE_HANDED_DAGGER", 
    ], 
    "Two-Handed Axe": ["TWO_HANDED_AXE", ], 
    "Two-Handed Mace": ["TWO_HANDED_MACE", ], 
    "Two-Handed Sword": ["TWO_HANDED_SPEAR", ], 
    "Two-Handed Spear": ["TWO_HANDED_STAFF", ], 
    "Two-Handed Staff": ["TWO_HANDED_SWORD", ], 
    "Bow": ["BOW", ], 
    "All Two-Handed Weapons": [
        "TWO_HANDED_AXE", "TWO_HANDED_MACE", "TWO_HANDED_SPEAR", "TWO_HANDED_STAFF", "TWO_HANDED_SWORD", "BOW", 
    ], 
    "All Weapons": [
        "ONE_HANDED_AXE", "ONE_HANDED_MACES", "ONE_HANDED_SCEPTRE", "ONE_HANDED_SWORD", "WAND", "ONE_HANDED_DAGGER", 
        "TWO_HANDED_AXE", "TWO_HANDED_MACE", "TWO_HANDED_SPEAR", "TWO_HANDED_STAFF", "TWO_HANDED_SWORD", "BOW", 
    ], 
    "Off-Hand Catalyst": ["CATALYST", ], 
    "Shield": ["SHIELD", ], 
    "Quiver": ["QUIVER", ], 
    "All Off-Hand": [
        "CATALYST", "SHIELD", "QUIVER", 
    ], 
    "Helmet": ["HELMET", ], 
    "Body Armor": ["BODY_ARMOR", ], 
    "Belt": ["BELT", ], 
    "Boots": ["BOOTS", ], 
    "Gloves": ["GLOVES", ], 
    "All Armour": [
        "HELMET", "BODY_ARMOR", "BELT", "BOOTS", "GLOVES", 
    ], 
    "Amulet": ["AMULET", ], 
    "Ring": ["RING", ], 
    "Relic": ["RELIC", ], 
    "All Jewellery": [
        "AMULET", "RING", "RELIC", 
    ], 
    "All Equipment": [
        "CATALYST", "SHIELD", "QUIVER", 
        "HELMET", "BODY_ARMOR", "BELT", "BOOTS", "GLOVES", 
        "AMULET", "RING", "RELIC", 
    ], 
    "Small Idol": ["IDOL_1x1_ETERRA", ], 
    "Minor Idol": ["IDOL_1x1_LAGON", ], 
    "Humble Idol": ["IDOL_2x1", ], 
    "Stout Idol": ["IDOL_1x2", ], 
    "All Generic Idols": [
        "IDOL_1x1_ETERRA", "IDOL_1x1_LAGON", "IDOL_2x1", "IDOL_1x2", 
    ], 
    "Grand Idol": ["IDOL_3x1", ], 
    "Huge Idol": ["IDOL_1x3", ], 
    "Ornate Idol": ["IDOL_4x1", ], 
    "Large Idol": ["IDOL_1x4", ], 
    "Adorned Idol": ["IDOL_2x2", ], 
    "All Class-Specific Idols": [
        "IDOL_3x1", "IDOL_1x3", "IDOL_4x1", "IDOL_1x4", "IDOL_2x2", 
    ], 
    "All Idols": [
        "IDOL_1x1_ETERRA", "IDOL_1x1_LAGON", "IDOL_2x1", "IDOL_1x2", 
        "IDOL_3x1", "IDOL_1x3", "IDOL_4x1", "IDOL_1x4", "IDOL_2x2", 
    ], 
}

filepaths = dict()
filepaths["project"] = Path(".").absolute()
filepaths = {
    "project": filepaths["project"], 
    "filter_maker": filepaths["project"].joinpath("filter_maker"), 
    "affix_id": filepaths["project"].joinpath("affix_id"), 
    "item_id": filepaths["project"].joinpath("item_id"), 
    "unique_id": filepaths["project"].joinpath("unique_id"), 
    "condition_list": filepaths["project"].joinpath("condition_list"), 
}
filepaths["filter_maker"] = {
    "raw": filepaths["filter_maker"].joinpath("raw"), 
    "input": filepaths["filter_maker"].joinpath("input"), 
    "output": filepaths["filter_maker"].joinpath("output"), 
}
for key in ["affix_id", "item_id", "unique_id"]:
    filepaths[key] = {
        "combined": filepaths[key].joinpath("combined"), 
        "game_data": {
            "processed": filepaths[key].joinpath("game_data", "processed"), 
            "raw": filepaths[key].joinpath("game_data", "raw"), 
        }, 
        "details": {
            "processed": filepaths[key].joinpath("details", "processed"), 
            "raw": filepaths[key].joinpath("details", "raw"), 
        }, 
    }

def strip_ws(text: str) -> str:
    stripped_text = re.compile(r"\s+").sub(" ", text)
    stripped_text = stripped_text.replace(u"\u00d7", "x")
    stripped_text = stripped_text.replace(u"\u2013", "-")
    stripped_text = stripped_text.replace("( ", "(")
    stripped_text = stripped_text.replace(" )", ")")
    stripped_text = stripped_text.replace(" ,", ",")
    return stripped_text

def replace_if_na(val: Union[float, int, str, bool], rep: Union[float, int, str, bool]) -> Union[float, int, str, bool]:
    if val == val:
        return val
    else:
        return rep

def get_text_width_pixels(text: str, font_name: str = "calibri", font_size: int = 11) -> float:
    """
    Calculate the actual pixel width of text using PIL, accounting for font characteristics.
    
    Args:
        text: The text to measure
        font_name: Font name (calibri, arial, times, etc.)
        font_size: Font size in points
        
    Returns:
        Text width in pixels
    """
    try:
        # Create a dummy image for text measurement
        img = Image.new('RGB', (1, 1))
        draw = ImageDraw.Draw(img)
        
        """
        # Try to load the font - common font paths for different OS
        font_paths = [
            f"/mnt/c/Windows/Fonts/{font_name}.ttf",  # Windows
            f"/mnt/c/Windows/Fonts/{font_name}i.ttf",  # Windows italic variant
            f"/System/Library/Fonts/{font_name}.ttf",  # macOS
            f"/usr/share/fonts/truetype/{font_name}.ttf",  # Linux
        ]
        
        font = None
        for path in font_paths:
            if os.path.exists(path):
                try:
                    font = ImageFont.truetype(path, font_size)
                    break
                except:
                    continue
        
        # Fall back to default font if specific font not found
        if font is None:
            font = ImageFont.load_default()
        """
        
        font = ImageFont.load_default(size=font_size)
        
        # Get text bounding box
        bbox = draw.textbbox((0, 0), str(text), font=font)
        width_pixels = bbox[2] - bbox[0]
        
        return width_pixels
        
    except Exception:
        # Fallback to character count method if PIL fails
        return len(str(text)) * 7  # Approximate pixels per character

def calculate_wrapped_lines(text: str, font_name: str, font_size: float, column_width_pixels: float) -> int:
    """
    Calculate how many lines text will wrap to in a given column width.
   
    Args:
        text: The text content
        font_name: Font name (e.g., "calibri")
        font_size: Font size in points
        column_width_pixels: Available width in pixels
   
    Returns:
        Number of lines the text will wrap to
    """
    if not text:
        return 1
   
    # Handle explicit line breaks
    lines = text.split('\n')
    total_lines = 0
   
    for line in lines:
        if not line:
            total_lines += 1
            continue
           
        # Calculate width of this line
        line_width_pixels = get_text_width_pixels(line, font_name, font_size)
       
        # Calculate how many visual lines this logical line will take
        if line_width_pixels <= column_width_pixels:
            total_lines += 1
        else:
            # Estimate wrapped lines - this is approximate
            wrapped_lines = max(1, int(line_width_pixels / column_width_pixels) + 1)
            total_lines += wrapped_lines
   
    return max(1, total_lines)

def pixels_to_excel_width(pixels: float) -> float:
    """
    Convert pixel width to Excel column width units.
    Excel width units are based on the width of '0' in the default font.
    
    Args:
        pixels: Width in pixels
        
    Returns:
        Width in Excel units
    """
    # Excel width unit is approximately 7 pixels for default font (Calibri 11pt)
    # This is a reasonable approximation
    return pixels / 6.5

def excel_width_to_pixels(excel_width: float) -> float:
    """
    Convert Excel column width units to pixels.
    Excel width is based on the width of '0' in the default font.
   
    Args:
        excel_width: Width in Excel units
       
    Returns:
        Width in pixels
    """
    # Approximate conversion: Excel width * 7 pixels per unit
    # This is based on default Calibri 11pt font
    return excel_width * 7.5

def format_sheet(
    sheet: Worksheet,
    h_alignment: str = "left",
    v_alignment: str = "top", 
    freeze_panes: Tuple[int, int] = (1, 1),
    auto_width: bool = True,
    wrap_text: bool = False
) -> None:
    """
    Format a single worksheet with alignment, freeze panes, auto-width columns, and auto-height rows.
    
    Args:
        sheet: openpyxl Worksheet object
        h_alignment: Horizontal alignment - "left", "centre", or "right"
        v_alignment: Vertical alignment - "top", "centre", or "bottom" 
        freeze_panes: Tuple of (row, column) to freeze at
        auto_width: Whether to automatically adjust column widths
        wrap_text: Whether to enable text wrapping in cells
    """
    # Map alignment strings to openpyxl constants
    h_map = {"left": "left", "centre": "center", "right": "right"}
    v_map = {"top": "top", "centre": "center", "bottom": "bottom"}
    
    # Create alignment object
    alignment = Alignment(
        horizontal=h_map.get(h_alignment, "left"),
        vertical=v_map.get(v_alignment, "top"),
        wrap_text=wrap_text
    )
    
    # Apply alignment to all cells with data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = alignment
    
    # Set freeze panes
    row, col = freeze_panes
    if row > 0 and col > 0:
        # Convert to cell reference (e.g., (2,2) -> "B2")
        freeze_cell = sheet.cell(row=row, column=col).coordinate
        sheet.freeze_panes = freeze_cell
   
    # Store column widths for row height calculation
    column_widths = {}
    
    # Auto-adjust column widths using PIL for accurate text measurement
    if auto_width:
        for column in sheet.columns:
            max_width_pixels = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # Get the font info from the cell (if available)
                        font_name = "calibri"  # Default Excel font
                        font_size = 11  # Default Excel font size
                        
                        # Try to get actual font from cell formatting
                        if cell.font and cell.font.name:
                            font_name = cell.font.name.lower()
                        if cell.font and cell.font.size:
                            font_size = cell.font.size
                        
                        # Calculate actual text width in pixels
                        text_width_pixels = get_text_width_pixels(
                            str(cell.value), font_name, font_size
                        )
                        
                        if text_width_pixels > max_width_pixels:
                            max_width_pixels = text_width_pixels
                except:
                    pass
            
            # Convert pixels to Excel width units and add padding
            if max_width_pixels > 0:
                excel_width = pixels_to_excel_width(max_width_pixels) + 4  # Add padding
                # Cap maximum width at reasonable size
                excel_width = min(excel_width, 80)
                sheet.column_dimensions[column_letter].width = excel_width
                column_widths[column_letter] = excel_width
            else:
                # Store default width for empty columns
                column_widths[column_letter] = sheet.column_dimensions[column_letter].width or 8.43
    else:
        # If auto_width is disabled, get existing column widths
        for column in sheet.columns:
            column_letter = column[0].column_letter
            column_widths[column_letter] = sheet.column_dimensions[column_letter].width or 8.43
   
    # Auto-adjust row heights when wrap_text is enabled
    if wrap_text:
        for row_cells in sheet.iter_rows():
            max_height_needed = 15  # Default Excel row height in points
            row_number = row_cells[0].row
           
            for cell in row_cells:
                if cell.value is not None:
                    try:
                        # Get font info
                        font_name = "calibri"
                        font_size = 11
                       
                        if cell.font and cell.font.name:
                            font_name = cell.font.name.lower()
                        if cell.font and cell.font.size:
                            font_size = cell.font.size
                       
                        # Get column width in pixels
                        column_letter = cell.column_letter
                        column_width_excel = column_widths.get(column_letter, 8.43)
                        column_width_pixels = excel_width_to_pixels(column_width_excel)
                       
                        # Calculate how many lines the text will wrap to
                        text = str(cell.value)
                        lines_needed = calculate_wrapped_lines(
                            text, font_name, font_size, column_width_pixels
                        )
                       
                        # Calculate required height (font_size + padding per line)
                        line_height = font_size * 1.2  # Line spacing factor
                        required_height = lines_needed * line_height + 3  # Add padding
                       
                        if required_height > max_height_needed:
                            max_height_needed = required_height
                   
                    except:
                        pass
           
            # Set row height if it needs to be taller than default
            if max_height_needed > 15:
                sheet.row_dimensions[row_number].height = max_height_needed

def format_workbook(
    file: Union[str, Workbook], 
    h_alignment: str = "left", 
    v_alignment: str = "top", 
    freeze_panes: Tuple[int, int] = (1, 1),
    auto_width: bool = True,
    wrap_text: bool = False
) -> None:
    """
    Format all worksheets in a workbook with alignment, freeze panes, and auto-width.
    
    Args:
        file: Either a filepath (str) or an openpyxl Workbook object
        h_alignment: Horizontal alignment - "left", "centre", or "right" 
        v_alignment: Vertical alignment - "top", "centre", or "bottom"
        freeze_panes: Tuple of (row, column) to freeze at
        auto_width: Whether to automatically adjust column widths
    """
    # If file is a string, load the workbook and call recursively
    if isinstance(file, (str, Path, )):
        wb = load_workbook(file)
        format_workbook(wb, h_alignment, v_alignment, freeze_panes, auto_width, wrap_text)
        wb.save(file)  # Save changes back to file
        return
    
    # If file is a Workbook, format each sheet
    wb = file
    for sheet in wb.worksheets:
        format_sheet(sheet, h_alignment, v_alignment, freeze_panes, auto_width, wrap_text)

def parse_markdown_to_structured_data(file_path: str) -> list[dict]:
    """
    Parse markdown file and convert to structured data suitable for Excel
    """
    with open(file_path, "r", encoding="utf-8") as file:
        lines = file.readlines()
    
    structured_data = []
    
    for line in lines:
        line = line.rstrip("\n")
        
        # Skip empty lines
        if not line.strip():
            continue
            
        # Detect headers (# ## ###)
        if line.startswith("#"):
            level = len(line) - len(line.lstrip("#"))
            text = line.lstrip("#").strip()
            structured_data.append({
                "type": "header",
                "level": level,
                "text": text,
                "indent": 0
            })
        
        # Detect numbered lists (1. 2. etc.)
        elif re.match(r"^\s*\d+\.", line):
            indent_level = (len(line) - len(line.lstrip())) // 4  # Assuming 4 spaces per indent
            text = re.sub(r"^\s*\d+\.\s*", "", line)
            structured_data.append({
                "type": "numbered_list",
                "level": 1,
                "text": text,
                "indent": indent_level
            })
        
        # Detect unordered lists with letters (a. b. etc.)
        elif re.match(r"^\s*[a-z]\.\s", line):
            indent_level = (len(line) - len(line.lstrip())) // 4
            text = re.sub(r"^\s*[a-z]\.\s*", "", line)
            structured_data.append({
                "type": "lettered_list",
                "level": 2,
                "text": text,
                "indent": indent_level
            })
        
        # Regular text (including sub-points)
        else:
            indent_level = (len(line) - len(line.lstrip())) // 4
            text = line.strip()
            structured_data.append({
                "type": "text",
                "level": 0,
                "text": text,
                "indent": indent_level
            })
    
    return structured_data

def create_excel_from_structured_data(structured_data: list[dict], sheet_name: str, output_file: str) -> None:
    """
    Create Excel file with formatting from structured markdown data
    """
    # Create DataFrame
    df_data = []
    for item in structured_data:
        # Add indentation to text based on indent level
        indent_prefix = "    " * item["indent"]
        formatted_text = indent_prefix + item["text"]
        
        df_data.append({
            "Content": formatted_text,
            "Type": item["type"],
            "Level": item["level"]
        })
    
    df = pd.DataFrame(df_data)
    
    # Create workbook and write data
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write only the Content column to Excel (we"ll use Type and Level for formatting)
    for idx, item in enumerate(structured_data, 1):
        cell = ws[f"A{idx}"]
        
        # Add indentation to text
        indent_prefix = "    " * item["indent"]
        cell.value = indent_prefix + item["text"]
        
        # Apply formatting based on type
        if item["type"] == "header":
            if item["level"] == 1:
                cell.font = Font(size=16, bold=True)
            elif item["level"] == 2:
                cell.font = Font(size=14, bold=True)
            else:
                cell.font = Font(size=12, bold=True)
        
        elif item["type"] in ["numbered_list", "lettered_list"]:
            cell.font = Font(size=11, bold=True)
        
        else:  # regular text
            cell.font = Font(size=11)
        
        # Set alignment
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column width
    ws.column_dimensions["A"].width = 100
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved as: {output_file}")

def markdown_to_excel(markdown_file: str, excel_file: str) -> None:
    """
    Main function to convert markdown to Excel
    """
    print(f"Reading markdown file: {markdown_file}")
    structured_data = parse_markdown_to_structured_data(markdown_file)
    
    print(f"Converting to Excel: {excel_file}")
    create_excel_from_structured_data(structured_data, excel_file)
    
    print(f"Successfully converted {markdown_file} to {excel_file}")
